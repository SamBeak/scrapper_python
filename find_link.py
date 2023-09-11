from requests import get
from bs4 import BeautifulSoup
from selenium import webdriver

import json # json data
import openpyxl # excel data

# 엑셀 저장을 위한 선언
filename = "business_listings_anchor.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active

# JSON 파일을 위한 선언
business_list = []
with open('./python/business_list.json', 'r', encoding='utf-8') as f:
    business_list = json.load(f)

# Selenium을 위한 선언
browser = webdriver.Chrome()

base_url = "https://www.bizno.net/?query="

for item in business_list:
    for i in range(len(business_list[item])):
        search_term = business_list[item][i]["name"]
        ceo = business_list[item][i]["ceo"]
        location = business_list[item][i]["location"]
        print(f"{base_url}{search_term}")
        browser.get(f"{base_url}{search_term}")     # selenium 을 사용해서 scraping을 해오겠다.

        results = [] # 결과값 초기화
        soup = BeautifulSoup(browser.page_source, "html.parser")
        business_name_lists = soup.find_all("div", class_="details")
        banner = soup.find("div", class_="banner-content")
        h3 = banner.find("h3", class_="text-white")
        number = h3.find("strong").string
        if number != '0':
            for bs in business_name_lists:
                anchor = bs.find("a")
                location_name = bs.find("p").string
                if anchor != None and ceo != None:
                    h4 = anchor.find("h4")
                    title = h4.string
                    link = anchor["href"]
                    link = link.split("/")[2]
                    ceo_name = bs.find("h5")
                    ceo_name = ceo_name.string.split(":")[1].strip()
                    if ceo_name == ceo or location_name == location:
                        business_data = {
                            'title' : title,
                            'link' : link,
                            'ceo' : ceo_name,
                            'location' : location_name
                        }
                        results.append(business_data)
                        sheet.cell(row=i+1, column=1).value = business_data['title']
                        sheet.cell(row=i+1, column=2).value = business_data['link']
                        sheet.cell(row=i+1, column=3).value = business_data['ceo']
                        sheet.cell(row=i+1, column=4).value = business_data['location']
                    else:
                        pass
                else :
                    pass
wb.save(filename)